import streamlit as st
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import json, os, base64, zipfile
import requests
from datetime import date, datetime
from pypdf import PdfReader, PdfWriter
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from batch_generate import render_batch_tab

TEMPLATE_UNITED = "backflow_template.pdf"
TEMPLATE_JAX = "jacksonville_template.pdf"
TECHNICIANS_FILE = "technicians.json"
PAGE_W, PAGE_H = 612, 792
JAX_PAGE_W, JAX_PAGE_H = 612, 792  # safe default; overridden in main() if template exists

UNITED_STICKY_FIELDS = [
    "customer_name", "street_address", "branch", "ahj",
    "manufacturer", "model", "size",
]
JAX_STICKY_FIELDS = [
    "premises_name", "owner_name", "service_address", "mailing_address",
    "contact_phone", "jea_account", "meter_number",
    "manufacturer", "model_number", "size",
]

UNITED_RESET_FIELDS = [
    "serial_number", "location",
    "assembly_type", "system_service", "bypass",
    "cv1_result", "cv1_dp", "cv2_result", "cv2_dp",
    "rv_result", "rv_psi", "rv_out_result", "rv_in_result",
    "pvb_ai_result", "pvb_ai_psi", "pvb_cv_result", "pvb_cv_psi",
    "assembly_result", "repair_desc",
]
JAX_RESET_FIELDS = [
    "device_type", "serial_number", "install_date", "physical_location",
    "comm_test_purpose", "comm_service_type", "comm_reclaim", "comm_fire_bypass",
    "res_test_purpose", "res_service_type", "res_reclaim",
    "init_cv1_result", "init_cv1_psi", "init_cv2_result", "init_cv2_psi",
    "init_rv_result", "init_rv_psi", "init_pvb_result", "init_pvb_psi",
    "final_cv1_result", "final_cv1_psi", "final_cv2_result", "final_cv2_psi",
    "final_rv_result", "final_rv_psi", "final_pvb_result",
    "assembly_result", "repairs",
]

UNITED_RESET_WIDGET_KEYS = [
    "u_sn", "u_loc",
    "u_cv1dp", "u_cv2dp", "u_rvpsi", "u_aipsi", "u_cvpsi", "u_rep",
]
JAX_RESET_WIDGET_KEYS = [
    "j_dt", "j_sn", "j_id", "j_pl",
    "j_icv1p", "j_icv2p", "j_irvp", "j_ipvbp",
    "j_fcv1p", "j_fcv2p", "j_frvp", "j_rep",
]


def _get_pdf_page_size(path):
    try:
        reader = PdfReader(path)
        page = reader.pages[0]
        mb = page.mediabox
        return float(mb.width), float(mb.height)
    except Exception:
        return 595, 842


# ─────────────────────────────────────────
# Technician helpers
# ─────────────────────────────────────────

# ─────────────────────────────────────────
# GitHub-backed persistence for technicians.json
#
# Streamlit Cloud containers have EPHEMERAL local disk: every sleep/wake or
# redeploy re-clones the repo fresh, wiping any local-only file writes.
# To survive that, every technician save/delete also commits the updated
# technicians.json straight back to this GitHub repo via the REST API, so
# the next container start pulls the latest version instead of a blank one.
#
# Requires a Streamlit secret named GITHUB_TOKEN (a fine-grained GitHub
# Personal Access Token with "Contents: Read and write" on this repo).
# If the secret is missing, the app silently falls back to local-disk-only
# storage (previous behavior) so nothing breaks without it configured.
# ─────────────────────────────────────────
GITHUB_REPO = "kellyjwinkle/backflow_app"
GITHUB_BRANCH = "main"
GITHUB_API_URL = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{TECHNICIANS_FILE}"


def _github_token():
    try:
        return st.secrets.get("GITHUB_TOKEN", "")
    except Exception:
        return ""


def _github_headers():
    token = _github_token()
    if not token:
        return None
    return {"Authorization": f"token {token}", "Accept": "application/vnd.github+json"}


def _github_fetch_technicians():
    """Pull the latest technicians.json from GitHub. Returns (data, sha) or
    (None, None) if unavailable (no token, network error, file not found)."""
    headers = _github_headers()
    if not headers:
        return None, None
    try:
        resp = requests.get(f"{GITHUB_API_URL}?ref={GITHUB_BRANCH}", headers=headers, timeout=8)
        if resp.status_code != 200:
            return None, None
        payload = resp.json()
        content = base64.b64decode(payload["content"]).decode("utf-8")
        return json.loads(content), payload.get("sha")
    except Exception:
        return None, None


def _github_push_technicians(data: dict):
    """Push the current technicians dict to GitHub. Returns (ok, message)."""
    headers = _github_headers()
    if not headers:
        return False, "GITHUB_TOKEN not configured — saved locally only (won't survive a redeploy)."
    try:
        sha = st.session_state.get("_technicians_github_sha")
        if not sha:
            _, sha = _github_fetch_technicians()
        body = {
            "message": "Update technicians.json via app",
            "content": base64.b64encode(json.dumps(data, indent=2).encode("utf-8")).decode("utf-8"),
            "branch": GITHUB_BRANCH,
        }
        if sha:
            body["sha"] = sha
        resp = requests.put(GITHUB_API_URL, headers=headers, json=body, timeout=8)
        if resp.status_code in (200, 201):
            new_sha = resp.json().get("content", {}).get("sha")
            if new_sha:
                st.session_state["_technicians_github_sha"] = new_sha
            return True, "Synced to GitHub ✓"
        return False, f"GitHub sync failed ({resp.status_code}): {resp.text[:200]}"
    except Exception as e:
        return False, f"GitHub sync error: {e}"


def _init_technicians():
    if "technicians" not in st.session_state:
        gh_data, gh_sha = _github_fetch_technicians()
        if gh_data is not None:
            st.session_state["technicians"] = gh_data
            st.session_state["_technicians_github_sha"] = gh_sha
        elif os.path.exists(TECHNICIANS_FILE):
            with open(TECHNICIANS_FILE, "r") as fh:
                st.session_state["technicians"] = json.load(fh)
        else:
            st.session_state["technicians"] = {}


def get_technician_names():
    _init_technicians()
    return [""] + list(st.session_state["technicians"].keys())


def get_technician_profile(name: str) -> dict:
    _init_technicians()
    return dict(st.session_state["technicians"].get(name, {}))


def upsert_technician_profile(name: str, profile: dict):
    _init_technicians()
    st.session_state["technicians"][name] = profile
    try:
        with open(TECHNICIANS_FILE, "w") as fh:
            json.dump(st.session_state["technicians"], fh, indent=2)
        local_ok, local_msg = True, "Profile saved ✓"
    except Exception as e:
        local_ok, local_msg = False, f"Save error: {e}"

    gh_ok, gh_msg = _github_push_technicians(st.session_state["technicians"])
    if local_ok and gh_ok:
        return True, "Profile saved ✓ (synced to GitHub)"
    if local_ok and not gh_ok:
        return True, f"Profile saved locally, but {gh_msg}"
    return False, local_msg


def delete_technician_profile(name: str):
    _init_technicians()
    if not name or name not in st.session_state["technicians"]:
        return False, "Profile not found."
    del st.session_state["technicians"][name]
    try:
        with open(TECHNICIANS_FILE, "w") as fh:
            json.dump(st.session_state["technicians"], fh, indent=2)
        local_ok, local_msg = True, "Profile deleted."
    except Exception as e:
        local_ok, local_msg = False, str(e)

    gh_ok, gh_msg = _github_push_technicians(st.session_state["technicians"])
    if local_ok and gh_ok:
        return True, "Profile deleted (synced to GitHub)."
    if local_ok and not gh_ok:
        return True, f"Profile deleted locally, but {gh_msg}"
    return False, local_msg


# ─────────────────────────────────────────
# In-session job store (PDFs in memory / direct device download)
# ─────────────────────────────────────────

def _jobs_store():
    if "_session_jobs" not in st.session_state:
        st.session_state["_session_jobs"] = []
    return st.session_state["_session_jobs"]


def add_job_to_session(form_data: dict, pdf_bytes: bytes, form_type: str):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S%f")
    serial = str(form_data.get("serial_number", "")).replace(" ", "_").replace("/", "-")
    if form_type == "united":
        cust = form_data.get("customer_name", "unknown").replace(" ", "_").replace("/", "-")
        base = f"united_{cust}_{serial}_{ts}" if serial else f"united_{cust}_{ts}"
    else:
        prem = form_data.get("premises_name", "unknown").replace(" ", "_").replace("/", "-")
        base = f"jax_{prem}_{serial}_{ts}" if serial else f"jax_{prem}_{ts}"

    existing_filenames = {j.get("filename") for j in _jobs_store()}
    filename = f"{base}.pdf"
    suffix = 1
    while filename in existing_filenames:
        filename = f"{base}_{suffix}.pdf"
        suffix += 1

    job_entry = {
        "filename": filename,
        "form_type": form_type,
        "saved_at": datetime.now().strftime("%m/%d/%Y %H:%M"),
        "saved_date": datetime.now().strftime("%Y-%m-%d"),
        "technician": form_data.get("technician", ""),
        "date": form_data.get("date") or form_data.get("signature_date", ""),
        "customer": form_data.get("customer_name") or form_data.get("premises_name", ""),
        "address": form_data.get("street_address") or form_data.get("service_address", ""),
        "serial_number": form_data.get("serial_number", ""),
        "location": form_data.get("location") or form_data.get("physical_location", ""),
        "assembly_type": form_data.get("assembly_type") or form_data.get("device_type", ""),
        "assembly_result": form_data.get("assembly_result", ""),
        "pdf_bytes": pdf_bytes,
    }
    _jobs_store().append(job_entry)
    return filename


def build_jobs_excel(jobs: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Summary"
    headers = [
        "Saved At", "Form Type", "Date", "Technician", "Customer / Premises",
        "Address", "Location / Building", "Serial Number", "Assembly Type", "Result", "PDF Filename",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, job in enumerate(jobs, 2):
        ws.cell(row=row_idx, column=1, value=job.get("saved_at", ""))
        ws.cell(row=row_idx, column=2, value=job.get("form_type", "").upper())
        ws.cell(row=row_idx, column=3, value=job.get("date", ""))
        ws.cell(row=row_idx, column=4, value=job.get("technician", ""))
        ws.cell(row=row_idx, column=5, value=job.get("customer", ""))
        ws.cell(row=row_idx, column=6, value=job.get("address", ""))
        ws.cell(row=row_idx, column=7, value=job.get("location", ""))
        ws.cell(row=row_idx, column=8, value=job.get("serial_number", ""))
        ws.cell(row=row_idx, column=9, value=job.get("assembly_type", ""))
        ws.cell(row=row_idx, column=10, value=job.get("assembly_result", ""))
        ws.cell(row=row_idx, column=11, value=job.get("filename", ""))
        result_cell = ws.cell(row=row_idx, column=10)
        if job.get("assembly_result") == "PASSED":
            result_cell.fill = PatternFill("solid", fgColor="C6EFCE")
            result_cell.font = Font(color="276221")
        elif job.get("assembly_result") == "FAILED":
            result_cell.fill = PatternFill("solid", fgColor="FFC7CE")
            result_cell.font = Font(color="9C0006")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_session_zip(selected_jobs: list, all_jobs: list) -> bytes:
    today_str = datetime.now().strftime("%Y-%m-%d")
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        excel_bytes = build_jobs_excel(all_jobs)
        zf.writestr(f"backflow_summary_{today_str}.xlsx", excel_bytes)
        for job in selected_jobs:
            pdf_bytes = job.get("pdf_bytes")
            if pdf_bytes:
                zf.writestr(job["filename"], pdf_bytes)
    buf.seek(0)
    return buf.getvalue()


def reset_form_for_new_job(form_type: str):
    if form_type == "united":
        form = st.session_state.get("united_form", {})
        for f in UNITED_RESET_FIELDS:
            form.pop(f, None)
        _clear_widget_keys(UNITED_RESET_WIDGET_KEYS)
        for k in ["u_assembly_type", "u_system_service", "u_bypass",
                  "u_cv1_result", "u_cv2_result", "u_rv_result",
                  "u_rv_out_result", "u_rv_in_result",
                  "u_pvb_ai_result", "u_pvb_cv_result", "u_assembly_result"]:
            st.session_state.pop(k, None)
    else:
        form = st.session_state.get("jax_form", {})
        for f in JAX_RESET_FIELDS:
            form.pop(f, None)
        _clear_widget_keys(JAX_RESET_WIDGET_KEYS)
        for k in ["j_dt", "j_comm_test_purpose", "j_comm_service_type", "j_comm_reclaim",
                  "j_res_test_purpose", "j_res_service_type", "j_res_reclaim",
                  "j_init_cv1_result", "j_init_cv2_result", "j_init_rv_result",
                  "j_init_pvb_result", "j_final_cv1_result", "j_final_cv2_result",
                  "j_final_rv_result", "j_final_pvb_result", "j_assembly_result"]:
            st.session_state.pop(k, None)


UNITED_TEXT_FIELDS = {
    "date": (135, 583, 8), "branch": (235, 583, 8), "ahj": (437, 583, 8),
    "customer_name": (200, 567, 8), "street_address": (200, 551, 8), "location": (200, 533, 8),
    "serial_number": (205, 507, 8), "manufacturer": (205, 490, 8), "model": (205, 475, 8), "size": (390, 507, 8),
    "rv_psi": (300, 398, 8), "cv1_dp": (183, 320, 8), "cv2_dp": (395, 312, 8), "pvb_ai_psi": (495, 378, 8),
    "pvb_cv_psi": (495, 320, 8), "test_date": (168, 290, 8), "gauge_mfg": (215, 178, 8),
    "gauge_serial": (313, 178, 8), "date_cal": (455, 178, 8), "technician": (176, 155, 8),
    "cert_no": (407, 165, 8), "recert": (407, 150, 8),
}
UNITED_SIG_X, UNITED_SIG_Y, UNITED_SIG_W, UNITED_SIG_H = 170, 118, 130, 28
UNITED_CHECKBOXES = {
    "RP": (210,460), "DC": (270,460), "PVB": (210,441), "SVB": (270,441), "FIRE": (395,490), "DOMESTIC": (395,475),
    "IRRIGATION": (395,460), "ATTRACTION": (395,442), "BYPASS_YES": (500,470), "BYPASS_NO": (500,450),
    "CV1_CLOSED": (130,390), "CV1_LEAKED": (130,375), "CV2_CLOSED": (330,390), "CV2_LEAKED": (330,375),
    "PVB_AI_CLOSED": (426,398), "PVB_AI_OPENED": (426,378), "PVB_CV_LEAKED": (426,350), "PVB_CV_HELD": (426,323),
    "RV_OPENED": (225,398), "RV_DIDNOTOPEN": (225,378), "RV_OUT_CLOSED": (225,334), "RV_OUT_LEAKED": (272,334),
    "RV_IN_CLOSED": (225,310), "RV_IN_LEAKED": (272,310), "PASSED": (360,292), "FAILED": (415,292),
}
UNITED_REPAIR_BOX = (228, 200, 10, 3, 70)

JAX_TEXT_FIELDS = {
    "premises_name": (102, 696, 9), "owner_name": (333, 696, 9), "service_address": (104, 652, 9),
    "mailing_address": (335, 652, 9), "physical_location": (105, 614, 9), "contact_phone": (333, 612, 9),
    "jea_account": (103, 569, 9), "meter_number": (335, 571, 9), "device_type": (90, 419, 9),
    "manufacturer": (151, 421, 9), "size": (231, 421, 9), "model_number": (283, 420, 9),
    "serial_number": (358, 420, 9), "install_date": (459, 418, 9), "init_cv1_psi": (151, 338, 9),
    "init_cv2_psi": (249, 338, 9), "init_rv_psi": (396, 356, 9), "init_pvb_psi": (471, 335, 9),
    "final_cv1_psi": (165, 282, 9), "final_cv2_psi": (263, 283, 9), "final_rv_psi": (404, 299, 9),
    "repairs": (99, 244, 9), "init_tester_name": (94, 182, 9), "init_company": (236, 183, 9),
    "init_cert": (356, 183, 9), "init_test_date": (464, 183, 9), "repaired_by": (95, 158, 9),
    "repair_company": (239, 156, 9), "repair_cert": (355, 159, 9), "repair_date": (464, 162, 9),
    "final_tester_name": (93, 135, 9), "final_company": (239, 135, 9), "final_cert": (354, 136, 9),
    "final_test_date": (468, 139, 9), "signature_date": (433, 84, 9),
}
JAX_SIG_X, JAX_SIG_Y, JAX_SIG_W, JAX_SIG_H = 161, 68, 160, 22
JAX_CHECKBOXES = {
    "COMM_ANNUAL": (214, 545), "COMM_REPAIR": (286, 544), "COMM_REPLACEMENT": (358, 545), "COMM_NEW_INSTALL": (463, 545),
    "COMM_FIRE": (214, 523), "COMM_IRRIGATION": (294, 522), "COMM_PROCESS": (362, 521), "COMM_POTABLE": (472, 523),
    "COMM_FIRE_BYPASS": (215, 510), "RECLAIM_YES": (421, 511), "RECLAIM_NO": (459, 510), "RES_ANNUAL": (210, 489),
    "RES_REPAIR": (280, 488), "RES_REPLACEMENT": (358, 489), "RES_NEW_INSTALL": (462, 489), "RES_POTABLE": (202, 466),
    "RES_IRRIGATION": (255, 465), "RES_RECLAIM_YES": (434, 466), "RES_RECLAIM_NO": (472, 464), "INIT_CV1_CLOSED": (139, 363),
    "INIT_CV2_CLOSED": (235, 362), "INIT_RV_OPENED": (331, 356), "INIT_RV_DIDNOT": (336, 329), "INIT_PVB_AIOPEN": (445, 359),
    "INIT_PVB_AIDNOT": (451, 323), "FINAL_CV1_CLOSED": (138, 306), "FINAL_CV2_CLOSED": (236, 301),
    "FINAL_RV_OPENED": (331, 296), "FINAL_PVB_SAT": (450, 290), "JAX_PASSED": (300, 106), "JAX_FAILED": (358, 108),
}

UNITED_TESTER_DISPLAY_KEYS = [
    "u_gmfg_display", "u_gsn_display", "u_cal_display",
    "u_tech_display", "u_cert_display", "u_recert_display",
]
JAX_TESTER_DISPLAY_KEYS = [
    "j_itn_display", "j_ico_display", "j_ic_display",
    "j_rb_display", "j_rco_display", "j_rc_display",
    "j_ftn_display", "j_fco_display", "j_fc_display",
]
UNITED_TESTER_WIDGET_KEYS = ["u_gmfg", "u_gsn", "u_cal", "u_tech", "u_cert", "u_recert"]
JAX_TESTER_WIDGET_KEYS = ["j_itn", "j_ico", "j_ic", "j_rb", "j_rco", "j_rc", "j_ftn", "j_fco", "j_fc"]
TESTER_KEYS = ["gauge_mfg", "gauge_serial", "date_cal", "technician", "cert_no", "recert"]
JAX_TESTER_MAP = {
    "init_tester_name": "technician", "init_company": "company", "init_cert": "cert_no",
    "repaired_by": "technician", "repair_company": "company", "repair_cert": "cert_no",
    "final_tester_name": "technician", "final_company": "company", "final_cert": "cert_no",
}


def draw_x(c, bx, by, size=3.8):
    c.setStrokeColorRGB(1, 0, 0)
    c.setLineWidth(2.0)
    c.line(bx-size, by-size, bx+size, by+size)
    c.line(bx+size, by-size, bx-size, by+size)


def put_text(c, val, x, y, sz=8):
    if val:
        c.setFont("Helvetica-Bold", sz)
        c.setFillColorRGB(1, 0, 0)
        c.drawString(x, y, str(val))


def wrap_text(text, w=58):
    words = text.split()
    lines, line = [], ""
    for word in words:
        t = (line + " " + word).strip()
        if len(t) > w:
            lines.append(line.strip())
            line = word
        else:
            line = t
    if line:
        lines.append(line)
    return lines


def _init_form(key, defaults=None):
    if key not in st.session_state:
        st.session_state[key] = defaults or {}


def _clear_widget_keys(keys):
    for key in keys:
        st.session_state.pop(key, None)


def clearable_input(label, form_key, field_key, widget_key, **kwargs):
    form = st.session_state[form_key]
    if widget_key not in st.session_state:
        st.session_state[widget_key] = form.get(field_key, "")
    val = st.text_input(label, key=widget_key, **kwargs)
    form[field_key] = val
    return val


def get_signature_image_reader(form_data=None):
    source_b64 = form_data.get("signature_b64") if form_data and form_data.get("signature_b64") else st.session_state.get("signature_b64")
    if source_b64:
        try:
            buf = BytesIO(base64.b64decode(source_b64))
            buf.seek(0)
            return ImageReader(buf)
        except Exception:
            return None
    return None


def clear_signature():
    st.session_state.pop("signature_b64", None)


def save_signature(img_array):
    buf = BytesIO()
    img = Image.fromarray(img_array.astype("uint8"), "RGBA")
    img.save(buf, format="PNG")
    b64 = base64.b64encode(buf.getvalue()).decode()
    st.session_state["signature_b64"] = b64
    return b64


def _date_to_string(value):
    if not value:
        return ""
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    return str(value)


def synced_date_input(label, form_key, source_key, widget_key, target_fields):
    form = st.session_state[form_key]
    current = form.get(source_key, "")
    default_date = date.today()
    if current:
        for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
            try:
                default_date = datetime.strptime(str(current), fmt).date()
                break
            except Exception:
                pass
    picked = st.date_input(label, value=default_date, key=widget_key)
    picked_str = _date_to_string(picked)
    form[source_key] = picked_str
    for field in target_fields:
        form[field] = picked_str
    return picked


def apply_profile_to_forms(profile: dict):
    """Write a technician profile into both forms and clear stale widget keys."""
    _init_form("united_form")
    _init_form("jax_form")

    if not profile:
        _clear_widget_keys(
            UNITED_TESTER_WIDGET_KEYS + JAX_TESTER_WIDGET_KEYS
            + UNITED_TESTER_DISPLAY_KEYS + JAX_TESTER_DISPLAY_KEYS
        )
        return

    if profile.get("signature_b64"):
        st.session_state["signature_b64"] = profile["signature_b64"]

    united = st.session_state["united_form"]
    for tk in TESTER_KEYS:
        united[tk] = profile.get(tk, "")
    united["signature_b64"] = profile.get("signature_b64", "")

    jax = st.session_state["jax_form"]
    for jk, pk in JAX_TESTER_MAP.items():
        jax[jk] = profile.get(pk, "")
    for tk in TESTER_KEYS:
        jax[tk] = profile.get(tk, "")
    jax["signature_b64"] = profile.get("signature_b64", "")

    _clear_widget_keys(
        UNITED_TESTER_WIDGET_KEYS + JAX_TESTER_WIDGET_KEYS
        + UNITED_TESTER_DISPLAY_KEYS + JAX_TESTER_DISPLAY_KEYS
    )


def generate_united_pdf(form_data: dict) -> bytes:
    reader = PdfReader(TEMPLATE_UNITED)
    writer = PdfWriter()
    writer.append(reader)
    packet = BytesIO()
    c = canvas.Canvas(packet, pagesize=(PAGE_W, PAGE_H))
    for field, (x, y, sz) in UNITED_TEXT_FIELDS.items():
        put_text(c, form_data.get(field, ""), x, y, sz)
    atype = form_data.get("assembly_type", "")
    if atype == "RP": draw_x(c, *UNITED_CHECKBOXES["RP"])
    elif atype == "DC": draw_x(c, *UNITED_CHECKBOXES["DC"])
    elif atype == "PVB": draw_x(c, *UNITED_CHECKBOXES["PVB"])
    elif atype == "SVB": draw_x(c, *UNITED_CHECKBOXES["SVB"])
    svc = form_data.get("system_service", "")
    if svc == "Fire": draw_x(c, *UNITED_CHECKBOXES["FIRE"])
    elif svc == "Domestic": draw_x(c, *UNITED_CHECKBOXES["DOMESTIC"])
    elif svc == "Irrigation": draw_x(c, *UNITED_CHECKBOXES["IRRIGATION"])
    elif svc == "Attraction": draw_x(c, *UNITED_CHECKBOXES["ATTRACTION"])
    bypass = form_data.get("bypass", "")
    if bypass == "Yes": draw_x(c, *UNITED_CHECKBOXES["BYPASS_YES"])
    elif bypass == "No": draw_x(c, *UNITED_CHECKBOXES["BYPASS_NO"])
    cv1r = form_data.get("cv1_result", "")
    if cv1r == "Closed Tight": draw_x(c, *UNITED_CHECKBOXES["CV1_CLOSED"])
    elif cv1r == "Leaked": draw_x(c, *UNITED_CHECKBOXES["CV1_LEAKED"])
    cv2r = form_data.get("cv2_result", "")
    if cv2r == "Closed Tight": draw_x(c, *UNITED_CHECKBOXES["CV2_CLOSED"])
    elif cv2r == "Leaked": draw_x(c, *UNITED_CHECKBOXES["CV2_LEAKED"])
    rvr = form_data.get("rv_result", "")
    if rvr == "Opened": draw_x(c, *UNITED_CHECKBOXES["RV_OPENED"])
    elif rvr == "Did Not Open": draw_x(c, *UNITED_CHECKBOXES["RV_DIDNOTOPEN"])
    rvo = form_data.get("rv_out_result", "")
    if rvo == "Closed Tight": draw_x(c, *UNITED_CHECKBOXES["RV_OUT_CLOSED"])
    elif rvo == "Leaked": draw_x(c, *UNITED_CHECKBOXES["RV_OUT_LEAKED"])
    rvi = form_data.get("rv_in_result", "")
    if rvi == "Closed Tight": draw_x(c, *UNITED_CHECKBOXES["RV_IN_CLOSED"])
    elif rvi == "Leaked": draw_x(c, *UNITED_CHECKBOXES["RV_IN_LEAKED"])
    pai = form_data.get("pvb_ai_result", "")
    if pai == "Opened": draw_x(c, *UNITED_CHECKBOXES["PVB_AI_OPENED"])
    elif pai == "Did Not Open": draw_x(c, *UNITED_CHECKBOXES["PVB_AI_CLOSED"])
    pcv = form_data.get("pvb_cv_result", "")
    if pcv == "Leaked": draw_x(c, *UNITED_CHECKBOXES["PVB_CV_LEAKED"])
    elif pcv == "Held": draw_x(c, *UNITED_CHECKBOXES["PVB_CV_HELD"])
    ar = form_data.get("assembly_result", "")
    if ar == "PASSED": draw_x(c, *UNITED_CHECKBOXES["PASSED"])
    elif ar == "FAILED": draw_x(c, *UNITED_CHECKBOXES["FAILED"])
    repair = form_data.get("repair_desc", "")
    if repair:
        bx, by, bh, bl, bw = UNITED_REPAIR_BOX
        lines = wrap_text(repair, bw)
        c.setFont("Helvetica-Bold", 7)
        c.setFillColorRGB(1, 0, 0)
        for i, ln in enumerate(lines[:bl]):
            c.drawString(bx, by - i * bh, ln)
    sig_reader = get_signature_image_reader(form_data)
    if sig_reader:
        c.drawImage(sig_reader, UNITED_SIG_X, UNITED_SIG_Y, width=UNITED_SIG_W, height=UNITED_SIG_H, mask="auto")
    c.save()
    packet.seek(0)
    overlay_reader = PdfReader(packet)
    page = writer.pages[0]
    page.merge_page(overlay_reader.pages[0])
    out = BytesIO()
    writer.write(out)
    return out.getvalue()


def generate_jax_pdf(form_data: dict) -> bytes:
    pw = globals().get("JAX_PAGE_W", 612)
    ph = globals().get("JAX_PAGE_H", 792)
    reader = PdfReader(TEMPLATE_JAX)
    writer = PdfWriter()
    writer.append(reader)
    packet = BytesIO()
    c = canvas.Canvas(packet, pagesize=(pw, ph))
    for field, (x, y, sz) in JAX_TEXT_FIELDS.items():
        put_text(c, form_data.get(field, ""), x, y, sz)
    comm_tp = form_data.get("comm_test_purpose", "")
    if comm_tp == "Annual": draw_x(c, *JAX_CHECKBOXES["COMM_ANNUAL"])
    elif comm_tp == "Repair": draw_x(c, *JAX_CHECKBOXES["COMM_REPAIR"])
    elif comm_tp == "Replacement": draw_x(c, *JAX_CHECKBOXES["COMM_REPLACEMENT"])
    elif comm_tp == "New Install": draw_x(c, *JAX_CHECKBOXES["COMM_NEW_INSTALL"])
    comm_st = form_data.get("comm_service_type", "")
    if comm_st == "Fire": draw_x(c, *JAX_CHECKBOXES["COMM_FIRE"])
    elif comm_st == "Irrigation": draw_x(c, *JAX_CHECKBOXES["COMM_IRRIGATION"])
    elif comm_st == "Process": draw_x(c, *JAX_CHECKBOXES["COMM_PROCESS"])
    elif comm_st == "Potable": draw_x(c, *JAX_CHECKBOXES["COMM_POTABLE"])
    comm_fp = form_data.get("comm_fire_bypass", "")
    if comm_fp: draw_x(c, *JAX_CHECKBOXES["COMM_FIRE_BYPASS"])
    comm_rc = form_data.get("comm_reclaim", "")
    if comm_rc == "Yes": draw_x(c, *JAX_CHECKBOXES["RECLAIM_YES"])
    elif comm_rc == "No": draw_x(c, *JAX_CHECKBOXES["RECLAIM_NO"])
    res_tp = form_data.get("res_test_purpose", "")
    if res_tp == "Annual": draw_x(c, *JAX_CHECKBOXES["RES_ANNUAL"])
    elif res_tp == "Repair": draw_x(c, *JAX_CHECKBOXES["RES_REPAIR"])
    elif res_tp == "Replacement": draw_x(c, *JAX_CHECKBOXES["RES_REPLACEMENT"])
    elif res_tp == "New Install": draw_x(c, *JAX_CHECKBOXES["RES_NEW_INSTALL"])
    res_st = form_data.get("res_service_type", "")
    if res_st == "Potable": draw_x(c, *JAX_CHECKBOXES["RES_POTABLE"])
    elif res_st == "Irrigation": draw_x(c, *JAX_CHECKBOXES["RES_IRRIGATION"])
    res_rc = form_data.get("res_reclaim", "")
    if res_rc == "Yes": draw_x(c, *JAX_CHECKBOXES["RES_RECLAIM_YES"])
    elif res_rc == "No": draw_x(c, *JAX_CHECKBOXES["RES_RECLAIM_NO"])
    icv1 = form_data.get("init_cv1_result", "")
    if icv1 == "Closed Tight": draw_x(c, *JAX_CHECKBOXES["INIT_CV1_CLOSED"])
    icv2 = form_data.get("init_cv2_result", "")
    if icv2 == "Closed Tight": draw_x(c, *JAX_CHECKBOXES["INIT_CV2_CLOSED"])
    irv = form_data.get("init_rv_result", "")
    if irv == "Opened": draw_x(c, *JAX_CHECKBOXES["INIT_RV_OPENED"])
    elif irv == "Did Not Open": draw_x(c, *JAX_CHECKBOXES["INIT_RV_DIDNOT"])
    ipvb = form_data.get("init_pvb_result", "")
    if ipvb == "Air Inlet Opened": draw_x(c, *JAX_CHECKBOXES["INIT_PVB_AIOPEN"])
    elif ipvb == "Air Inlet Did Not": draw_x(c, *JAX_CHECKBOXES["INIT_PVB_AIDNOT"])
    fcv1 = form_data.get("final_cv1_result", "")
    if fcv1 == "Closed Tight": draw_x(c, *JAX_CHECKBOXES["FINAL_CV1_CLOSED"])
    fcv2 = form_data.get("final_cv2_result", "")
    if fcv2 == "Closed Tight": draw_x(c, *JAX_CHECKBOXES["FINAL_CV2_CLOSED"])
    frv = form_data.get("final_rv_result", "")
    if frv == "Opened": draw_x(c, *JAX_CHECKBOXES["FINAL_RV_OPENED"])
    fpvb = form_data.get("final_pvb_result", "")
    if fpvb == "Satisfactory": draw_x(c, *JAX_CHECKBOXES["FINAL_PVB_SAT"])
    ar = form_data.get("assembly_result", "")
    if ar == "PASSED": draw_x(c, *JAX_CHECKBOXES["JAX_PASSED"])
    elif ar == "FAILED": draw_x(c, *JAX_CHECKBOXES["JAX_FAILED"])
    sig_reader = get_signature_image_reader(form_data)
    if sig_reader:
        c.drawImage(sig_reader, JAX_SIG_X, JAX_SIG_Y, width=JAX_SIG_W, height=JAX_SIG_H, mask="auto")
    c.save()
    packet.seek(0)
    overlay_reader = PdfReader(packet)
    page = writer.pages[0]
    page.merge_page(overlay_reader.pages[0])
    out = BytesIO()
    writer.write(out)
    return out.getvalue()


def render_technician_sidebar():
    st.sidebar.title("👤 Technician")
    names = get_technician_names()
    prev_sel = st.session_state.get("_sidebar_tech_sel", "")
    if prev_sel not in names:
        prev_sel = ""

    selected = st.sidebar.selectbox(
        "Active Profile",
        names,
        index=names.index(prev_sel) if prev_sel in names else 0,
        key="sidebar_tech_select",
    )
    st.session_state["_sidebar_tech_sel"] = selected

    last_loaded = st.session_state.get("_last_loaded_tech", None)
    if selected and selected != last_loaded:
        profile = get_technician_profile(selected)
        apply_profile_to_forms(profile)
        st.session_state["_last_loaded_tech"] = selected
        st.rerun()

    if selected:
        sig_b64 = st.session_state.get("signature_b64")
        if sig_b64:
            st.sidebar.image(base64.b64decode(sig_b64), caption="Signature on file", use_container_width=True)
        else:
            st.sidebar.caption("⚠️ No signature for this profile.")

    st.sidebar.divider()

    with st.sidebar.expander("✏️ Signature", expanded=False):
        upload = st.file_uploader("Upload (PNG preferred)", type=["png", "jpg", "jpeg"], key="sig_upload")
        if upload is not None:
            try:
                img = Image.open(upload).convert("RGBA")
                buf = BytesIO()
                img.save(buf, format="PNG")
                st.session_state["signature_b64"] = base64.b64encode(buf.getvalue()).decode()
                st.success("Signature uploaded.")
            except Exception as e:
                st.error(f"Could not read signature: {e}")

        st.caption("Or draw:")
        try:
            from streamlit_drawable_canvas import st_canvas
            canvas_result = st_canvas(
                fill_color="rgba(255,255,255,0)", stroke_width=2, stroke_color="#000000",
                background_color="#FFFFFF", height=80, width=220,
                drawing_mode="freedraw", key="sb_sig_canvas",
            )
            if canvas_result.image_data is not None:
                arr = canvas_result.image_data
                if arr.max() > 0 and arr[:, :, 3].max() > 0:
                    save_signature(arr)
        except ImportError:
            st.info("Install streamlit-drawable-canvas for drawing.")

        col_sig1, col_sig2 = st.columns(2)
        with col_sig1:
            if st.button("🗑️ Clear", key="sb_clear_sig"):
                clear_signature()
                st.rerun()
        with col_sig2:
            if selected and st.button("💾 Save to Profile", key="save_sig_to_profile"):
                current = get_technician_profile(selected)
                current["signature_b64"] = st.session_state.get("signature_b64", "")
                ok, msg = upsert_technician_profile(selected, current)
                if ok:
                    st.success(msg)
                else:
                    st.warning(msg)

    with st.sidebar.expander("⚙️ Edit / Add Profile", expanded=False):
        current = get_technician_profile(selected) if selected else {}
        prof_name = st.text_input("Name", value=selected, key="prof_name")
        prof_co = st.text_input("Company", value=current.get("company", ""), key="prof_co")
        prof_cert = st.text_input("Cert No.", value=current.get("cert_no", ""), key="prof_cert")
        prof_rec = st.text_input("Re-Cert", value=current.get("recert", ""), key="prof_rec")
        prof_gmfg = st.text_input("Gauge Mfg", value=current.get("gauge_mfg", ""), key="prof_gmfg")
        prof_gsn = st.text_input("Gauge SN", value=current.get("gauge_serial", ""), key="prof_gsn")
        prof_cal = st.text_input("Date Cal", value=current.get("date_cal", ""), key="prof_cal")
        if current.get("signature_b64"):
            st.caption("Saved signature:")
            st.image(base64.b64decode(current["signature_b64"]), width=180)
        if st.button("💾 Save Profile", key="save_profile_btn"):
            if prof_name.strip():
                new_profile = {
                    "technician": prof_name.strip(),
                    "company": prof_co.strip(),
                    "cert_no": prof_cert.strip(),
                    "recert": prof_rec.strip(),
                    "gauge_mfg": prof_gmfg.strip(),
                    "gauge_serial": prof_gsn.strip(),
                    "date_cal": prof_cal.strip(),
                    "signature_b64": st.session_state.get("signature_b64", current.get("signature_b64", "")),
                }
                ok, msg = upsert_technician_profile(prof_name.strip(), new_profile)
                if ok:
                    st.success(msg)
                    apply_profile_to_forms(new_profile)
                    st.session_state["_sidebar_tech_sel"] = prof_name.strip()
                    st.session_state["_last_loaded_tech"] = prof_name.strip()
                else:
                    st.warning(msg)
                st.rerun()
            else:
                st.error("Name cannot be empty.")

    if selected:
        confirm_key = "_confirm_delete_profile"
        if not st.session_state.get(confirm_key, False):
            if st.sidebar.button("🗑️ Delete Profile", key="delete_profile_btn"):
                st.session_state[confirm_key] = True
                st.rerun()
        else:
            st.sidebar.warning(f"Delete **{selected}**?")
            col_yes, col_no = st.sidebar.columns(2)
            with col_yes:
                if st.button("✅ Yes", key="confirm_delete_yes"):
                    ok, msg = delete_technician_profile(selected)
                    st.session_state[confirm_key] = False
                    if ok:
                        st.session_state["_sidebar_tech_sel"] = ""
                        st.session_state["_last_loaded_tech"] = None
                        clear_signature()
                        st.sidebar.success(msg)
                    else:
                        st.sidebar.warning(msg)
                    st.rerun()
            with col_no:
                if st.button("❌ Cancel", key="confirm_delete_no"):
                    st.session_state[confirm_key] = False
                    st.rerun()


def render_tester_banner(form_key: str, form_type: str):
    _init_form(form_key)
    form = st.session_state[form_key]
    selected_tech = st.session_state.get("_sidebar_tech_sel", "")

    if not selected_tech:
        st.warning("⚠️ Select a technician profile in the sidebar.")
        return

    with st.expander(f"🔧 Tester: {selected_tech} — tap to review/edit", expanded=False):
        reset_key = f"{form_type}_reset_tester"
        if st.button("🔄 Reset from Profile", key=reset_key):
            profile = get_technician_profile(selected_tech)
            if form_type == "united":
                for tk in TESTER_KEYS:
                    form[tk] = profile.get(tk, "")
                _clear_widget_keys(UNITED_TESTER_DISPLAY_KEYS)
            else:
                for jk, pk in JAX_TESTER_MAP.items():
                    form[jk] = profile.get(pk, "")
                _clear_widget_keys(JAX_TESTER_DISPLAY_KEYS)
            st.rerun()

        if form_type == "united":
            col1, col2 = st.columns(2)
            with col1:
                v = st.text_input("Gauge Mfg", value=form.get("gauge_mfg", ""), key="u_gmfg_display")
                form["gauge_mfg"] = v
                v = st.text_input("Gauge Serial", value=form.get("gauge_serial", ""), key="u_gsn_display")
                form["gauge_serial"] = v
                v = st.text_input("Date Calibrated", value=form.get("date_cal", ""), key="u_cal_display")
                form["date_cal"] = v
            with col2:
                v = st.text_input("Technician", value=form.get("technician", ""), key="u_tech_display")
                form["technician"] = v
                v = st.text_input("Cert No.", value=form.get("cert_no", ""), key="u_cert_display")
                form["cert_no"] = v
                v = st.text_input("Re-Cert Date", value=form.get("recert", ""), key="u_recert_display")
                form["recert"] = v
        else:
            st.markdown("**Initial Tester**")
            col1, col2, col3 = st.columns(3)
            with col1:
                v = st.text_input("Init Tester", value=form.get("init_tester_name", ""), key="j_itn_display")
                form["init_tester_name"] = v
            with col2:
                v = st.text_input("Init Company", value=form.get("init_company", ""), key="j_ico_display")
                form["init_company"] = v
            with col3:
                v = st.text_input("Init Cert", value=form.get("init_cert", ""), key="j_ic_display")
                form["init_cert"] = v

            assembly_result = form.get("assembly_result", "")
            if assembly_result == "FAILED":
                st.markdown("**Repaired By**")
                col1, col2, col3 = st.columns(3)
                with col1:
                    v = st.text_input("Repaired By", value=form.get("repaired_by", ""), key="j_rb_display")
                    form["repaired_by"] = v
                with col2:
                    v = st.text_input("Repair Company", value=form.get("repair_company", ""), key="j_rco_display")
                    form["repair_company"] = v
                with col3:
                    v = st.text_input("Repair Cert", value=form.get("repair_cert", ""), key="j_rc_display")
                    form["repair_cert"] = v
            else:
                for fk in ("repaired_by", "repair_company", "repair_cert"):
                    form.pop(fk, None)
                st.caption("ℹ️ Repair row appears only when Assembly Result is FAILED.")

            st.markdown("**Final Tester**")
            col1, col2, col3 = st.columns(3)
            with col1:
                v = st.text_input("Final Tester", value=form.get("final_tester_name", ""), key="j_ftn_display")
                form["final_tester_name"] = v
            with col2:
                v = st.text_input("Final Company", value=form.get("final_company", ""), key="j_fco_display")
                form["final_company"] = v
            with col3:
                v = st.text_input("Final Cert", value=form.get("final_cert", ""), key="j_fc_display")
                form["final_cert"] = v

        sig_b64 = st.session_state.get("signature_b64") or form.get("signature_b64")
        if sig_b64:
            st.image(base64.b64decode(sig_b64), caption="Signature on file", width=180)
        else:
            st.caption("⚠️ No signature. Upload in sidebar.")


def render_united_form():
    _init_form("united_form")
    form = st.session_state["united_form"]

    st.subheader("📋 United Fire Protection — Backflow Test Report")

    with st.expander("🏢 Site / Customer Info (sticky — stays between buildings)", expanded=not bool(form.get("customer_name"))):
        st.caption("These fields persist until you manually clear them. Edit freely.")
        col1, col2 = st.columns(2)
        with col1:
            clearable_input("Customer Name", "united_form", "customer_name", "u_cust")
            clearable_input("Street Address", "united_form", "street_address", "u_addr")
            clearable_input("Branch", "united_form", "branch", "u_branch")
        with col2:
            clearable_input("AHJ", "united_form", "ahj", "u_ahj")
            clearable_input("Manufacturer", "united_form", "manufacturer", "u_mfg")
            clearable_input("Model", "united_form", "model", "u_model")
            clearable_input("Size", "united_form", "size", "u_size")

    st.divider()
    st.markdown("**🔩 Assembly / Building Info** *(clears after each save)*")
    col1, col2 = st.columns(2)
    with col1:
        synced_date_input("Test Date", "united_form", "date", "u_date", ["test_date"])
        clearable_input("Serial Number", "united_form", "serial_number", "u_sn")
        clearable_input("Location / Building", "united_form", "location", "u_loc",
                        placeholder="e.g. Building 3 / Unit 101")
    with col2:
        atype_opts = ["", "RP", "DC", "PVB", "SVB"]
        atype = st.selectbox("Assembly Type", atype_opts,
                             index=atype_opts.index(form.get("assembly_type", "")) if form.get("assembly_type", "") in atype_opts else 0,
                             key="u_assembly_type")
        form["assembly_type"] = atype
        svc_opts = ["", "Fire", "Domestic", "Irrigation", "Attraction"]
        svc = st.selectbox("System Service", svc_opts,
                           index=svc_opts.index(form.get("system_service", "")) if form.get("system_service", "") in svc_opts else 0,
                           key="u_system_service")
        form["system_service"] = svc
        bypass_opts = ["", "Yes", "No"]
        bypass = st.selectbox("Bypass", bypass_opts,
                              index=bypass_opts.index(form.get("bypass", "")) if form.get("bypass", "") in bypass_opts else 0,
                              key="u_bypass")
        form["bypass"] = bypass

    st.divider()
    st.markdown("**📊 Test Results**")
    atype_val = form.get("assembly_type", "")

    if atype_val in ("", "RP", "DC"):
        col1, col2 = st.columns(2)
        ct_opts = ["", "Closed Tight", "Leaked"]
        with col1:
            cv1 = st.selectbox("CV1 Result", ct_opts,
                               index=ct_opts.index(form.get("cv1_result", "")) if form.get("cv1_result", "") in ct_opts else 0,
                               key="u_cv1_result")
            form["cv1_result"] = cv1
            cv1dp = st.text_input("CV1 Differential Pressure", value=form.get("cv1_dp", ""), key="u_cv1dp")
            form["cv1_dp"] = cv1dp
        with col2:
            cv2 = st.selectbox("CV2 Result", ct_opts,
                               index=ct_opts.index(form.get("cv2_result", "")) if form.get("cv2_result", "") in ct_opts else 0,
                               key="u_cv2_result")
            form["cv2_result"] = cv2
            cv2dp = st.text_input("CV2 Differential Pressure", value=form.get("cv2_dp", ""), key="u_cv2dp")
            form["cv2_dp"] = cv2dp

    if atype_val in ("", "RP"):
        st.markdown("*Relief Valve*")
        col1, col2, col3 = st.columns(3)
        rv_opts = ["", "Opened", "Did Not Open"]
        rvo_opts = ["", "Closed Tight", "Leaked"]
        with col1:
            rv = st.selectbox("RV Result", rv_opts,
                              index=rv_opts.index(form.get("rv_result", "")) if form.get("rv_result", "") in rv_opts else 0,
                              key="u_rv_result")
            form["rv_result"] = rv
        with col2:
            rvpsi = st.text_input("RV Opened At (psi)", value=form.get("rv_psi", ""), key="u_rvpsi")
            form["rv_psi"] = rvpsi
        with col3:
            rvo = st.selectbox("RV Outlet", rvo_opts,
                               index=rvo_opts.index(form.get("rv_out_result", "")) if form.get("rv_out_result", "") in rvo_opts else 0,
                               key="u_rv_out_result")
            form["rv_out_result"] = rvo
            rvi = st.selectbox("RV Inlet", rvo_opts,
                               index=rvo_opts.index(form.get("rv_in_result", "")) if form.get("rv_in_result", "") in rvo_opts else 0,
                               key="u_rv_in_result")
            form["rv_in_result"] = rvi

    if atype_val in ("", "PVB", "SVB"):
        st.markdown("*Pressure Vacuum Breaker*")
        col1, col2 = st.columns(2)
        pai_opts = ["", "Opened", "Did Not Open"]
        pcv_opts = ["", "Held", "Leaked"]
        with col1:
            pai = st.selectbox("Air Inlet Result", pai_opts,
                               index=pai_opts.index(form.get("pvb_ai_result", "")) if form.get("pvb_ai_result", "") in pai_opts else 0,
                               key="u_pvb_ai_result")
            form["pvb_ai_result"] = pai
            aipsi = st.text_input("Air Inlet (psi)", value=form.get("pvb_ai_psi", ""), key="u_aipsi")
            form["pvb_ai_psi"] = aipsi
        with col2:
            pcv = st.selectbox("CV Result", pcv_opts,
                               index=pcv_opts.index(form.get("pvb_cv_result", "")) if form.get("pvb_cv_result", "") in pcv_opts else 0,
                               key="u_pvb_cv_result")
            form["pvb_cv_result"] = pcv
            cvpsi = st.text_input("CV (psi)", value=form.get("pvb_cv_psi", ""), key="u_cvpsi")
            form["pvb_cv_psi"] = cvpsi

    st.divider()
    col1, col2 = st.columns(2)
    ar_opts = ["", "PASSED", "FAILED"]
    with col1:
        ar = st.selectbox("Assembly Result", ar_opts,
                          index=ar_opts.index(form.get("assembly_result", "")) if form.get("assembly_result", "") in ar_opts else 0,
                          key="u_assembly_result")
        form["assembly_result"] = ar
    with col2:
        rep = st.text_input("Repair Description", value=form.get("repair_desc", ""), key="u_rep")
        form["repair_desc"] = rep

    render_tester_banner("united_form", "united")

    st.divider()
    col_save, col_dl = st.columns(2)
    with col_save:
        if st.button("✅ Save & Next Building", key="u_save_job", type="primary", use_container_width=True):
            try:
                pdf_bytes = generate_united_pdf(form)
                filename = add_job_to_session(form, pdf_bytes, "united")
                st.success(f"✓ Saved: {filename}")
                reset_form_for_new_job("united")
                st.session_state.pop("u_pdf_bytes", None)
                st.rerun()
            except Exception as e:
                st.error(f"Save failed: {e}")

    if "u_pdf_bytes" not in st.session_state:
        if st.button("🖨️ Preview PDF (no save)", key="u_gen_pdf", use_container_width=True):
            try:
                st.session_state["u_pdf_bytes"] = generate_united_pdf(form)
            except Exception as e:
                st.error(f"PDF error: {e}")

    if "u_pdf_bytes" in st.session_state:
        with col_dl:
            loc_slug = form.get("location", "").replace(" ", "_")
            st.download_button(
                "⬇️ Download PDF",
                data=st.session_state["u_pdf_bytes"],
                file_name=f"backflow_{form.get('customer_name','report')}_{loc_slug}_{date.today().strftime('%Y%m%d')}.pdf",
                mime="application/pdf",
                key="u_dl_pdf",
                use_container_width=True,
            )


def render_jax_form():
    _init_form("jax_form")
    form = st.session_state["jax_form"]

    st.subheader("📋 Jacksonville — Backflow Test Report")

    with st.expander("🏢 Premises / Owner Info (sticky — stays between devices)", expanded=not bool(form.get("premises_name"))):
        st.caption("These fields persist until you manually clear them. Edit freely.")
        col1, col2 = st.columns(2)
        with col1:
            clearable_input("Premises Name", "jax_form", "premises_name", "j_pn")
            clearable_input("Service Address", "jax_form", "service_address", "j_sa")
            clearable_input("JEA Account #", "jax_form", "jea_account", "j_jea")
        with col2:
            clearable_input("Owner Name", "jax_form", "owner_name", "j_on")
            clearable_input("Mailing Address", "jax_form", "mailing_address", "j_ma")
            clearable_input("Contact Phone", "jax_form", "contact_phone", "j_cp")
            clearable_input("Meter Number", "jax_form", "meter_number", "j_mn")

    st.divider()
    st.markdown("**Commercial**")
    col1, col2 = st.columns(2)
    comm_tp_opts = ["", "Annual", "Repair", "Replacement", "New Install"]
    comm_st_opts = ["", "Fire", "Irrigation", "Process", "Potable"]
    comm_rc_opts = ["", "Yes", "No"]
    with col1:
        comm_tp = st.selectbox("Test Purpose", comm_tp_opts,
                               index=comm_tp_opts.index(form.get("comm_test_purpose", "")) if form.get("comm_test_purpose", "") in comm_tp_opts else 0,
                               key="j_comm_test_purpose")
        form["comm_test_purpose"] = comm_tp
        comm_st = st.selectbox("Service Type", comm_st_opts,
                               index=comm_st_opts.index(form.get("comm_service_type", "")) if form.get("comm_service_type", "") in comm_st_opts else 0,
                               key="j_comm_service_type")
        form["comm_service_type"] = comm_st
    with col2:
        comm_rc = st.selectbox("Reclaim", comm_rc_opts,
                               index=comm_rc_opts.index(form.get("comm_reclaim", "")) if form.get("comm_reclaim", "") in comm_rc_opts else 0,
                               key="j_comm_reclaim")
        form["comm_reclaim"] = comm_rc
        comm_fp = st.checkbox("Fire Service Bypass", value=bool(form.get("comm_fire_bypass", False)), key="j_comm_fire_bypass")
        form["comm_fire_bypass"] = comm_fp

    st.divider()
    st.markdown("**Residential**")
    col1, col2 = st.columns(2)
    res_tp_opts = ["", "Annual", "Repair", "Replacement", "New Install"]
    res_st_opts = ["", "Potable", "Irrigation"]
    res_rc_opts = ["", "Yes", "No"]
    with col1:
        res_tp = st.selectbox("Test Purpose (Res)", res_tp_opts,
                              index=res_tp_opts.index(form.get("res_test_purpose", "")) if form.get("res_test_purpose", "") in res_tp_opts else 0,
                              key="j_res_test_purpose")
        form["res_test_purpose"] = res_tp
        res_st = st.selectbox("Service Type (Res)", res_st_opts,
                              index=res_st_opts.index(form.get("res_service_type", "")) if form.get("res_service_type", "") in res_st_opts else 0,
                              key="j_res_service_type")
        form["res_service_type"] = res_st
    with col2:
        res_rc = st.selectbox("Reclaim (Res)", res_rc_opts,
                              index=res_rc_opts.index(form.get("res_reclaim", "")) if form.get("res_reclaim", "") in res_rc_opts else 0,
                              key="j_res_reclaim")
        form["res_reclaim"] = res_rc

    st.divider()
    st.markdown("**🔩 Device Info** *(clears after each save)*")
    col1, col2 = st.columns(2)
    with col1:
        dt_opts = ["", "RP", "DC", "PVB", "SVB"]
        dt = st.selectbox(
            "Device Type",
            dt_opts,
            index=dt_opts.index(form.get("device_type", "")) if form.get("device_type", "") in dt_opts else 0,
            key="j_dt",
        )
        form["device_type"] = dt
        clearable_input("Serial Number", "jax_form", "serial_number", "j_sn")
        clearable_input("Install Date", "jax_form", "install_date", "j_id")
        clearable_input("Physical Location", "jax_form", "physical_location", "j_pl",
                        placeholder="e.g. Building A / Meter Room")
    with col2:
        clearable_input("Manufacturer", "jax_form", "manufacturer", "j_jmfg")
        clearable_input("Model Number", "jax_form", "model_number", "j_jmod")
        clearable_input("Size", "jax_form", "size", "j_jsz")

    st.divider()
    st.markdown("**Initial Test Results**")
    col1, col2 = st.columns(2)
    ct_opts = ["", "Closed Tight", "Leaked"]
    rv_opts_j = ["", "Opened", "Did Not Open"]
    pvb_opts = ["", "Air Inlet Opened", "Air Inlet Did Not"]
    with col1:
        icv1 = st.selectbox("Init CV1", ct_opts,
                            index=ct_opts.index(form.get("init_cv1_result", "")) if form.get("init_cv1_result", "") in ct_opts else 0,
                            key="j_init_cv1_result")
        form["init_cv1_result"] = icv1
        icv1p = st.text_input("Init CV1 (psi)", value=form.get("init_cv1_psi", ""), key="j_icv1p")
        form["init_cv1_psi"] = icv1p
        icv2 = st.selectbox("Init CV2", ct_opts,
                            index=ct_opts.index(form.get("init_cv2_result", "")) if form.get("init_cv2_result", "") in ct_opts else 0,
                            key="j_init_cv2_result")
        form["init_cv2_result"] = icv2
        icv2p = st.text_input("Init CV2 (psi)", value=form.get("init_cv2_psi", ""), key="j_icv2p")
        form["init_cv2_psi"] = icv2p
    with col2:
        irv = st.selectbox("Init RV", rv_opts_j,
                           index=rv_opts_j.index(form.get("init_rv_result", "")) if form.get("init_rv_result", "") in rv_opts_j else 0,
                           key="j_init_rv_result")
        form["init_rv_result"] = irv
        irvp = st.text_input("Init RV (psi)", value=form.get("init_rv_psi", ""), key="j_irvp")
        form["init_rv_psi"] = irvp
        ipvb = st.selectbox("Init PVB", pvb_opts,
                            index=pvb_opts.index(form.get("init_pvb_result", "")) if form.get("init_pvb_result", "") in pvb_opts else 0,
                            key="j_init_pvb_result")
        form["init_pvb_result"] = ipvb
        ipvbp = st.text_input("Init PVB (psi)", value=form.get("init_pvb_psi", ""), key="j_ipvbp")
        form["init_pvb_psi"] = ipvbp

    st.divider()
    st.markdown("**Final Test Results**")
    col1, col2 = st.columns(2)
    frv_opts = ["", "Opened", "Did Not Open"]
    fpvb_opts = ["", "Satisfactory"]
    with col1:
        fcv1 = st.selectbox("Final CV1", ct_opts,
                            index=ct_opts.index(form.get("final_cv1_result", "")) if form.get("final_cv1_result", "") in ct_opts else 0,
                            key="j_final_cv1_result")
        form["final_cv1_result"] = fcv1
        fcv1p = st.text_input("Final CV1 (psi)", value=form.get("final_cv1_psi", ""), key="j_fcv1p")
        form["final_cv1_psi"] = fcv1p
        fcv2 = st.selectbox("Final CV2", ct_opts,
                            index=ct_opts.index(form.get("final_cv2_result", "")) if form.get("final_cv2_result", "") in ct_opts else 0,
                            key="j_final_cv2_result")
        form["final_cv2_result"] = fcv2
        fcv2p = st.text_input("Final CV2 (psi)", value=form.get("final_cv2_psi", ""), key="j_fcv2p")
        form["final_cv2_psi"] = fcv2p
    with col2:
        frv = st.selectbox("Final RV", frv_opts,
                           index=frv_opts.index(form.get("final_rv_result", "")) if form.get("final_rv_result", "") in frv_opts else 0,
                           key="j_final_rv_result")
        form["final_rv_result"] = frv
        frvp = st.text_input("Final RV (psi)", value=form.get("final_rv_psi", ""), key="j_frvp")
        form["final_rv_psi"] = frvp
        fpvb = st.selectbox("Final PVB", fpvb_opts,
                            index=fpvb_opts.index(form.get("final_pvb_result", "")) if form.get("final_pvb_result", "") in fpvb_opts else 0,
                            key="j_final_pvb_result")
        form["final_pvb_result"] = fpvb

    st.divider()
    col1, col2 = st.columns(2)
    ar_opts = ["", "PASSED", "FAILED"]
    with col1:
        ar = st.selectbox("Assembly Result", ar_opts,
                          index=ar_opts.index(form.get("assembly_result", "")) if form.get("assembly_result", "") in ar_opts else 0,
                          key="j_assembly_result")
        form["assembly_result"] = ar
    with col2:
        if form.get("assembly_result") == "FAILED":
            rep = st.text_input("Repairs Made", value=form.get("repairs", ""), key="j_rep")
            form["repairs"] = rep
        else:
            form.pop("repairs", None)
            st.caption("ℹ️ Repairs field appears only when Assembly Result is FAILED.")

    synced_date_input("Signature Date", "jax_form", "signature_date", "j_sigdate",
                      ["init_test_date", "final_test_date", "repair_date"])

    render_tester_banner("jax_form", "jax")

    st.divider()
    col_save, col_dl = st.columns(2)
    with col_save:
        if st.button("✅ Save & Next Building", key="j_save_job", type="primary", use_container_width=True):
            try:
                pdf_bytes = generate_jax_pdf(form)
                filename = add_job_to_session(form, pdf_bytes, "jax")
                st.success(f"✓ Saved: {filename}")
                reset_form_for_new_job("jax")
                st.session_state.pop("j_pdf_bytes", None)
                st.rerun()
            except Exception as e:
                st.error(f"Save failed: {e}")

    if "j_pdf_bytes" not in st.session_state:
        if st.button("🖨️ Preview PDF (no save)", key="j_gen_pdf", use_container_width=True):
            try:
                st.session_state["j_pdf_bytes"] = generate_jax_pdf(form)
            except Exception as e:
                st.error(f"PDF error: {e}")

    if "j_pdf_bytes" in st.session_state:
        with col_dl:
            st.download_button(
                "⬇️ Download PDF",
                data=st.session_state["j_pdf_bytes"],
                file_name=f"jax_backflow_{form.get('premises_name','report')}_{date.today().strftime('%Y%m%d')}.pdf",
                mime="application/pdf",
                key="j_dl_pdf",
                use_container_width=True,
            )


def render_jobs_tab():
    st.subheader("📁 Today's Jobs")
    all_jobs = _jobs_store()

    if not all_jobs:
        st.info("No jobs saved yet this session. Complete a form and tap 'Save & Next Building'.")
        return

    today_str = datetime.now().strftime("%Y-%m-%d")
    st.markdown(f"**{len(all_jobs)} report(s) this session — {datetime.now().strftime('%m/%d/%Y')}**")

    selected_filenames = []
    for idx, job in enumerate(all_jobs):
        fname = job.get("filename", "")
        result_icon = "✅" if job.get("assembly_result") == "PASSED" else ("❌" if job.get("assembly_result") == "FAILED" else "⬜")
        label = f"{result_icon}  {job.get('customer', '')} | {job.get('location', '')} | SN: {job.get('serial_number', '')} | {job.get('assembly_result', '')}"
        checked = st.checkbox(label, value=True, key=f"chk_{idx}_{fname}")
        if checked:
            selected_filenames.append(fname)
        pdf_bytes = job.get("pdf_bytes")
        if pdf_bytes:
            st.download_button(
                f"⬇️ {fname}",
                data=pdf_bytes,
                file_name=fname,
                mime="application/pdf",
                key=f"dl_{idx}_{fname}",
            )
        st.divider()

    st.divider()
    selected_jobs = [j for j in all_jobs if j.get("filename") in selected_filenames]
    if selected_jobs:
        col_zip, col_excel = st.columns(2)
        with col_zip:
            if st.button(f"📦 Build ZIP ({len(selected_jobs)} PDFs + Excel)", key="dl_zip", use_container_width=True):
                with st.spinner("Building ZIP..."):
                    zip_bytes = build_session_zip(selected_jobs, all_jobs)
                st.download_button(
                    "⬇️ Download ZIP",
                    data=zip_bytes,
                    file_name=f"backflow_jobs_{today_str}.zip",
                    mime="application/zip",
                    key="save_zip",
                    use_container_width=True,
                )
        with col_excel:
            if st.button("📊 Export Excel Only", key="export_excel", use_container_width=True):
                excel_bytes = build_jobs_excel(all_jobs)
                st.download_button(
                    "⬇️ Download Excel",
                    data=excel_bytes,
                    file_name=f"backflow_jobs_{today_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_excel",
                    use_container_width=True,
                )

    st.divider()
    confirm_clear = "_confirm_clear_session"
    if not st.session_state.get(confirm_clear, False):
        if st.button("🗑️ Clear Session Jobs", key="clear_session_btn"):
            st.session_state[confirm_clear] = True
            st.rerun()
    else:
        st.warning("⚠️ This will remove all jobs from this session.")
        col_yes, col_no = st.columns(2)
        with col_yes:
            if st.button("✅ Yes, clear", key="confirm_clear_yes"):
                st.session_state["_session_jobs"] = []
                st.session_state[confirm_clear] = False
                st.rerun()
        with col_no:
            if st.button("❌ Cancel", key="confirm_clear_no"):
                st.session_state[confirm_clear] = False
                st.rerun()


def main():
    st.set_page_config(
        page_title="Backflow Test Reports",
        page_icon="💧",
        layout="wide",
        initial_sidebar_state="collapsed",
    )

    if os.path.exists(TEMPLATE_JAX):
        global JAX_PAGE_W, JAX_PAGE_H
        JAX_PAGE_W, JAX_PAGE_H = _get_pdf_page_size(TEMPLATE_JAX)

    render_technician_sidebar()
    tab_united, tab_jax, tab_jobs, tab_batch = st.tabs(
        ["🔵 United Fire", "🟠 Jacksonville", "📁 Jobs", "📊 Batch Generate"]
    )
    with tab_united:
        render_united_form()
    with tab_jax:
        render_jax_form()
    with tab_jobs:
        render_jobs_tab()
    with tab_batch:
        render_batch_tab(generate_united_pdf, generate_jax_pdf, add_job_to_session)


if __name__ == "__main__":
    main()
